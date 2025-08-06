#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF Processor Core - Lógica de Processamento
============================================

Módulo central com toda a lógica de processamento de PDFs e atualização de Excel.
Separado da interface gráfica para melhor organização e reutilização.

Versão 3.2

Autor: Sistema de Extração Automatizada
Data: 2025
"""

import re
import pandas as pd
import pdfplumber
from datetime import datetime
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple, Callable
from openpyxl import load_workbook
import os
import shutil
from dotenv import load_dotenv

class PDFProcessorCore:
    """Classe central para processamento de PDFs - sem interface gráfica"""
    
    def __init__(self, progress_callback: Optional[Callable] = None, log_callback: Optional[Callable] = None):
        """
        Inicializa o processador
        
        Args:
            progress_callback: Função para atualizar progresso (0-100)
            log_callback: Função para enviar logs para interface
        """
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        
        # Regras de mapeamento para colunas específicas do Excel
        self.mapping_rules = {
            # FOLHA NORMAL - Obter da coluna ÍNDICE (com fallback especial para PRODUÇÃO)
            '01003601': {'code': 'PREMIO PROD. MENSAL', 'excel_column': 'X', 'source': 'indice', 'fallback_to_valor': True, 'folha_type': 'FOLHA NORMAL'},
            '01007301': {'code': 'HORAS EXT.100%-180', 'excel_column': 'Y', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '01009001': {'code': 'ADIC.NOT.25%-180', 'excel_column': 'AC', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '01003501': {'code': 'HORAS EXT.75%-180', 'excel_column': 'AA', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '02007501': {'code': 'DIFER.PROV. HORAS EXTRAS 75%', 'excel_column': 'AA', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            
            # FOLHA NORMAL - Obter da coluna VALOR
            '09090301_NORMAL': {'code': 'SALARIO CONTRIB INSS', 'excel_column': 'B', 'source': 'valor', 'folha_type': 'FOLHA NORMAL', 'original_code': '09090301'},
            
            # 13 SALARIO - Obter da coluna VALOR
            '09090301_13SAL': {'code': 'SALARIO CONTRIB INSS', 'excel_column': 'B', 'source': 'valor', 'folha_type': '13 SALARIO', 'original_code': '09090301'},
            '09090101_13SAL': {'code': 'REMUNERACAO BRUTA', 'excel_column': 'B', 'source': 'valor', 'folha_type': '13 SALARIO', 'original_code': '09090101', 'is_fallback': True}
        }
        
        # Planilha preferida
        self.preferred_sheet = None
        
        # Diretório de trabalho
        self.trabalho_dir = None
        
        # Meses em português para conversão
        self.meses_pt = {
            'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
            'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
            'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
        }
        
        self.meses_abrev = {
            'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
        }

    def _log(self, message: str, level: str = "INFO"):
        """Envia log para callback se disponível"""
        if self.log_callback:
            self.log_callback(f"[{level}] {message}")
        
        # Log padrão também
        if level == "DEBUG":
            logging.debug(message)
        elif level == "WARNING":
            logging.warning(message)
        elif level == "ERROR":
            logging.error(message)
        else:
            logging.info(message)

    def _update_progress(self, progress: int, message: str = ""):
        """Atualiza progresso se callback disponível"""
        if self.progress_callback:
            self.progress_callback(progress, message)

    def load_env_config(self):
        """Carrega configurações do arquivo .env"""
        if Path('.env').exists():
            load_dotenv()
            self._log("Arquivo .env carregado", "DEBUG")
        else:
            raise ValueError("Arquivo .env não encontrado. Configure MODELO_DIR no arquivo .env")
        
        self.trabalho_dir = os.getenv('MODELO_DIR')
        if not self.trabalho_dir:
            raise ValueError("MODELO_DIR não está definido no arquivo .env")
        
        trabalho_dir_path = Path(self.trabalho_dir)
        if not trabalho_dir_path.exists():
            raise ValueError(f"Diretório de trabalho não encontrado: {self.trabalho_dir}")
        
        self._log(f"Diretório de trabalho: {self.trabalho_dir}")

    def set_trabalho_dir(self, directory: str):
        """Define diretório de trabalho programaticamente"""
        if not os.path.exists(directory):
            raise ValueError(f"Diretório não encontrado: {directory}")
        
        self.trabalho_dir = directory
        self._log(f"Diretório de trabalho definido: {directory}")

    def validate_trabalho_dir(self) -> Tuple[bool, str]:
        """
        Valida se o diretório de trabalho está configurado e tem MODELO.xlsm
        
        Returns:
            Tuple[bool, str]: (válido, mensagem)
        """
        if not self.trabalho_dir:
            return False, "Diretório de trabalho não configurado"
        
        if not os.path.exists(self.trabalho_dir):
            return False, f"Diretório não encontrado: {self.trabalho_dir}"
        
        modelo_file = Path(self.trabalho_dir) / "MODELO.xlsm"
        if not modelo_file.exists():
            return False, f"MODELO.xlsm não encontrado em: {self.trabalho_dir}"
        
        return True, "Configuração válida"

    def get_pdf_files_in_trabalho_dir(self) -> List[str]:
        """Retorna lista de arquivos PDF no diretório de trabalho"""
        if not self.trabalho_dir or not os.path.exists(self.trabalho_dir):
            return []
        
        trabalho_path = Path(self.trabalho_dir)
        pdf_files = list(trabalho_path.glob("*.pdf"))
        return [pdf.name for pdf in pdf_files]

    def extract_person_name_from_pdf(self, pdf_path: str) -> Optional[str]:
        """Extrai o nome da pessoa da primeira página do PDF"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                if not pdf.pages:
                    return None
                
                first_page = pdf.pages[0]
                text = first_page.extract_text()
                
                if not text:
                    return None
                
                # Procura por padrões de nome
                name_patterns = [
                    r'Nome\s*:\s*([A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ\s]+?)(?:\n|$|[A-Z]{2,}:)',
                    r'NOME\s*:\s*([A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ\s]+?)(?:\n|$|[A-Z]{2,}:)',
                    r'Nome\s*:\s*(.+?)(?:\n|Endereço|CPF|RG)',
                    r'NOME\s*:\s*(.+?)(?:\n|ENDEREÇO|CPF|RG)',
                    r'Nome\s*:\s*(.+?)$',
                    r'NOME\s*:\s*(.+?)$'
                ]
                
                lines = text.split('\n')
                
                for line_num, line in enumerate(lines):
                    line_clean = line.strip()
                    
                    for pattern_num, pattern in enumerate(name_patterns):
                        match = re.search(pattern, line_clean, re.IGNORECASE)
                        if match:
                            nome_bruto = match.group(1).strip()
                            nome_limpo = self.clean_extracted_name(nome_bruto)
                            
                            if nome_limpo:
                                self._log(f"Nome detectado: {nome_limpo}", "DEBUG")
                                return nome_limpo
                
                return None
                
        except Exception as e:
            self._log(f"Erro ao extrair nome do PDF: {e}", "ERROR")
            return None

    def clean_extracted_name(self, nome_bruto: str) -> Optional[str]:
        """Limpa e valida o nome extraído"""
        if not nome_bruto:
            return None
        
        nome = nome_bruto.strip().upper()
        nome = re.sub(r'[^\w\s]', ' ', nome)
        nome = re.sub(r'\s+', ' ', nome).strip()
        
        if len(nome) < 3 or len(nome) > 100:
            return None
        
        if nome.replace(' ', '').isdigit():
            return None
        
        if not re.search(r'[A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ]', nome):
            return None
        
        palavras_excluir = ['NOME', 'FUNCIONARIO', 'FUNCIONÁRIO', 'TRABALHADOR', 'COLABORADOR', 'EMPREGADO']
        palavras = nome.split()
        palavras_filtradas = [p for p in palavras if p not in palavras_excluir]
        
        if not palavras_filtradas:
            return None
        
        nome_final = ' '.join(palavras_filtradas)
        
        if len(nome_final) < 3:
            return None
        
        return nome_final

    def normalize_filename(self, nome: str) -> str:
        """Converte nome da pessoa para formato de arquivo válido mantendo espaços"""
        filename = nome
        filename = re.sub(r'[<>:"/\\|?*]', '', filename)
        filename = re.sub(r'[\x00-\x1f\x7f]', '', filename)
        filename = re.sub(r'\s+', ' ', filename).strip()
        
        if len(filename) > 100:
            filename = filename[:100].rstrip()
        
        return filename

    def find_pdf_file(self, pdf_filename: str) -> str:
        """Procura arquivo PDF no diretório de trabalho"""
        if not self.trabalho_dir:
            raise ValueError("Diretório de trabalho não configurado")
        
        trabalho_dir_path = Path(self.trabalho_dir)
        
        # Se pdf_filename tem caminho completo
        if Path(pdf_filename).is_absolute() or '/' in pdf_filename or '\\' in pdf_filename:
            if Path(pdf_filename).exists():
                return pdf_filename
            else:
                raise ValueError(f"Arquivo PDF não encontrado: {pdf_filename}")
        
        # Procura no diretório de trabalho
        pdf_path = trabalho_dir_path / pdf_filename
        if pdf_path.exists():
            return str(pdf_path)
        
        # Tenta com extensão .pdf
        if not pdf_filename.lower().endswith('.pdf'):
            pdf_path_with_ext = trabalho_dir_path / f"{pdf_filename}.pdf"
            if pdf_path_with_ext.exists():
                return str(pdf_path_with_ext)
        
        raise ValueError(f"Arquivo PDF '{pdf_filename}' não encontrado no diretório de trabalho")

    def copy_modelo_to_dados(self, pdf_path: str, custom_name: Optional[str] = None) -> str:
        """Copia o arquivo modelo para a pasta DADOS"""
        trabalho_dir_path = Path(self.trabalho_dir)
        
        modelo_file = trabalho_dir_path / "MODELO.xlsm"
        if not modelo_file.exists():
            raise ValueError(f"Arquivo MODELO.xlsm não encontrado no diretório de trabalho")
        
        dados_dir = trabalho_dir_path / "DADOS"
        dados_dir.mkdir(exist_ok=True)
        
        if custom_name:
            base_name = self.normalize_filename(custom_name)
        else:
            base_name = Path(pdf_path).stem
        
        destino_file = dados_dir / f"{base_name}.xlsm"
        
        try:
            shutil.copy2(modelo_file, destino_file)
            return str(destino_file)
        except Exception as e:
            raise ValueError(f"Erro ao copiar modelo: {e}")

    def extract_text_from_pdf(self, pdf_path: str) -> List[str]:
        """Extrai texto de todas as páginas do PDF"""
        pages_text = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
                self._log(f"Processando PDF: {total_pages} páginas")
                
                for i, page in enumerate(pdf.pages):
                    # Atualiza progresso da extração (0-30% do total)
                    progress = int((i / total_pages) * 30)
                    self._update_progress(progress, f"Extraindo página {i+1}/{total_pages}")
                    
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
                    
        except Exception as e:
            self._log(f"Erro ao processar PDF: {e}", "ERROR")
            raise
            
        return pages_text

    def extract_reference_date(self, text: str) -> Optional[Tuple[int, int]]:
        """Extrai a data de referência da página (mês/ano)"""
        patterns = [
            r'Referência:\s*(\w+)/(\d{4})',
            r'Referencia:\s*(\w+)/(\d{4})',
            r'Data\s*do\s*c[aá]lculo:\s*\d{2}/(\d{2})/(\d{4})',
            r'Per[ií]odo:\s*(\w+)/(\d{4})',
            r'Compet[êe]ncia:\s*(\w+)/(\d{4})',
            r'(\w+)\s*/\s*(\d{4})',
        ]
        
        for i, pattern in enumerate(patterns):
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                try:
                    if len(match) == 2:
                        mes_str = match[0].lower()
                        ano = int(match[1])
                        
                        mes = self.meses_pt.get(mes_str) or self.meses_abrev.get(mes_str)
                        
                        if not mes:
                            try:
                                mes = int(mes_str)
                                if 1 <= mes <= 12:
                                    return (mes, ano)
                            except ValueError:
                                continue
                        else:
                            return (mes, ano)
                except ValueError:
                    continue
        
        return None

    def extract_last_two_numbers(self, line: str):
        """Extrai os dois últimos números de uma linha"""
        def convert_to_float_robust(value_str):
            if not value_str or not value_str.strip():
                return None
                
            cleaned = value_str.strip()
            
            # Detecta formato de horas (06:34) e converte para (06,34)
            if ':' in cleaned:
                hour_pattern = r'^\d{1,2}:\d{2}$'
                if re.match(hour_pattern, cleaned):
                    return cleaned.replace(':', ',')
            
            cleaned = re.sub(r'[^\d.,]', '', cleaned)
            
            if not cleaned:
                return None
            
            try:
                if ',' in cleaned and cleaned.count(',') == 1:
                    return float(cleaned.replace('.', '').replace(',', '.'))
                elif '.' in cleaned and cleaned.count('.') == 1 and ',' in cleaned:
                    return float(cleaned.replace(',', ''))
                elif ',' in cleaned and '.' not in cleaned:
                    return float(cleaned.replace(',', '.'))
                elif '.' in cleaned and ',' not in cleaned:
                    return float(cleaned)
                else:
                    return float(cleaned)
            except ValueError:
                return None
        
        number_pattern = r'[\d]+(?:[.,:]\d+)*'
        matches = re.findall(number_pattern, line)
        
        if len(matches) >= 2:
            penultimo = convert_to_float_robust(matches[-2])
            ultimo = convert_to_float_robust(matches[-1])
            return penultimo, ultimo
        elif len(matches) == 1:
            ultimo = convert_to_float_robust(matches[-1])
            return None, ultimo
        else:
            return None, None

    def extract_data_from_page(self, text: str, folha_type: str) -> Dict[str, any]:
        """Extrai dados específicos de uma página usando as regras de mapeamento"""
        data = {}
        codes_found = []
        
        relevant_rules = {k: v for k, v in self.mapping_rules.items() if v.get('folha_type') == folha_type}
        
        lines = text.split('\n')
        
        # Para 13 SALARIO, fallback especial entre 09090301 e 09090101
        found_09090301 = None
        found_09090101 = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            for rule_key, rule in relevant_rules.items():
                original_code = rule.get('original_code', rule_key)
                
                if original_code in line:
                    codes_found.append(original_code)
                    
                    indice, valor = self.extract_last_two_numbers(line)
                    
                    if folha_type == '13 SALARIO':
                        if original_code == '09090301':
                            found_09090301 = valor
                        elif original_code == '09090101':
                            found_09090101 = valor
                    
                    # Lógica normal para outros códigos
                    if folha_type == 'FOLHA NORMAL' or (folha_type == '13 SALARIO' and original_code not in ['09090301', '09090101']):
                        value_to_use = None
                        
                        if rule['source'] == 'indice':
                            if indice is not None and indice != 0:
                                value_to_use = indice
                            elif rule.get('fallback_to_valor', False) and valor is not None:
                                value_to_use = valor
                        elif rule['source'] == 'valor' and valor is not None:
                            value_to_use = valor
                        
                        if value_to_use is not None:
                            data[rule['excel_column']] = value_to_use
        
        # Para 13 SALARIO, aplica fallback
        if folha_type == '13 SALARIO':
            value_to_use = None
            
            if found_09090301 is not None and found_09090301 != 0:
                value_to_use = found_09090301
            elif found_09090101 is not None and found_09090101 != 0:
                value_to_use = found_09090101
            
            if value_to_use is not None:
                data['B'] = value_to_use
        
        return data

    def filter_and_categorize_pages(self, pages_text: List[str]) -> Dict[str, List[str]]:
        """Filtra e categoriza páginas por tipo"""
        categorized_pages = {
            'FOLHA NORMAL': [],
            '13 SALARIO': []
        }
        
        total_pages = len(pages_text)
        
        for i, text in enumerate(pages_text):
            # Atualiza progresso da categorização (30-40% do total)
            progress = int(30 + (i / total_pages) * 10)
            self._update_progress(progress, f"Categorizando página {i+1}/{total_pages}")
            
            page_type = None
            page_type_found = False
            
            lines = text.split('\n')
            for line in lines:
                line_clean = line.strip()
                
                if re.search(r'Tipo\s+da\s+folha\s*:', line_clean, re.IGNORECASE):
                    page_type_found = True
                    
                    if re.search(r'FOLHA\s+NORMAL', line_clean, re.IGNORECASE):
                        page_type = 'FOLHA NORMAL'
                        break
                    elif re.search(r'13\s*SAL[AÁ]RIO', line_clean, re.IGNORECASE):
                        page_type = '13 SALARIO'
                        break
                    elif re.search(r'F[ÉE]RIAS|ADIANTAMENTO|RESCIS[ÃA]O', line_clean, re.IGNORECASE):
                        page_type = 'IGNORAR'
                        break
            
            if not page_type_found:
                header_text = '\n'.join(lines[:10])
                
                if re.search(r'13\s*SAL[AÁ]RIO', header_text, re.IGNORECASE):
                    page_type = '13 SALARIO'
                elif re.search(r'F[ÉE]RIAS|ADIANTAMENTO\s*SALARIAL|RESCIS[ÃA]O', header_text, re.IGNORECASE):
                    page_type = 'IGNORAR'
                else:
                    page_type = 'FOLHA NORMAL'
            
            if page_type and page_type != 'IGNORAR':
                categorized_pages[page_type].append(text)
                
        self._log(f"Páginas categorizadas: FOLHA NORMAL={len(categorized_pages['FOLHA NORMAL'])}, 13 SALARIO={len(categorized_pages['13 SALARIO'])}")
        return categorized_pages

    def find_row_for_period(self, worksheet, month: int, year: int, folha_type: str) -> Optional[int]:
        """Encontra a linha correspondente ao período na planilha"""
        meses_nomes = ['', 'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                      'jul', 'ago', 'set', 'out', 'nov', 'dez']
        periodo_procurado = f"{meses_nomes[month]}/{str(year)[2:]}"
        
        if folha_type == 'FOLHA NORMAL':
            start_row = 1
            end_row = 65
        elif folha_type == '13 SALARIO':
            start_row = 67
            end_row = worksheet.max_row
        else:
            return None
        
        for row_num in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
            cell_value = worksheet[f'A{row_num}'].value
            
            if cell_value is None:
                continue
                
            if isinstance(cell_value, str):
                if cell_value.strip() == periodo_procurado:
                    return row_num
            elif isinstance(cell_value, datetime):
                if cell_value.month == month and cell_value.year == year:
                    return row_num
            elif isinstance(cell_value, (int, float)):
                try:
                    from datetime import timedelta
                    
                    if cell_value > 59:
                        excel_date = datetime(1899, 12, 30) + timedelta(days=cell_value)
                    else:
                        excel_date = datetime(1899, 12, 31) + timedelta(days=cell_value)
                    
                    if excel_date.month == month and excel_date.year == year:
                        return row_num
                        
                except Exception:
                    continue
        
        return None

    def update_excel_file(self, excel_path: str, extracted_data: Dict):
        """Atualiza o arquivo Excel existente com os dados extraídos"""
        try:
            is_macro_enabled = excel_path.lower().endswith('.xlsm')
            
            if is_macro_enabled:
                workbook = load_workbook(excel_path, keep_vba=True)
            else:
                workbook = load_workbook(excel_path)
            
            worksheet = None
            
            if self.preferred_sheet:
                if self.preferred_sheet in workbook.sheetnames:
                    worksheet = workbook[self.preferred_sheet]
                else:
                    raise ValueError(f"Planilha especificada '{self.preferred_sheet}' não encontrada")
            else:
                if 'LEVANTAMENTO DADOS' in workbook.sheetnames:
                    worksheet = workbook['LEVANTAMENTO DADOS']
                else:
                    raise ValueError("Planilha 'LEVANTAMENTO DADOS' não encontrada. Use -s para especificar outra planilha.")
            
            updates_count = 0
            failed_periods = []
            successful_periods = []
            
            total_periods = sum(len(periods) for periods in extracted_data.values())
            current_period = 0
            
            for folha_type in ['FOLHA NORMAL', '13 SALARIO']:
                if folha_type not in extracted_data:
                    continue
                    
                for (month, year), data in extracted_data[folha_type].items():
                    current_period += 1
                    
                    # Atualiza progresso do Excel (70-100% do total)
                    progress = int(70 + (current_period / total_periods) * 30)
                    meses = ['', 'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                            'jul', 'ago', 'set', 'out', 'nov', 'dez']
                    periodo = f"{meses[month]}/{str(year)[2:]}"
                    self._update_progress(progress, f"Atualizando {periodo} ({folha_type})")
                    
                    row_num = self.find_row_for_period(worksheet, month, year, folha_type)
                    
                    if row_num:
                        period_updates = 0
                        
                        for column, value in data.items():
                            cell_address = f"{column}{row_num}"
                            old_value = worksheet[cell_address].value
                            
                            if old_value is None or old_value == '' or old_value == 0:
                                worksheet[cell_address] = value
                                updates_count += 1
                                period_updates += 1
                        
                        if period_updates > 0:
                            successful_periods.append(f"{periodo} ({folha_type})")
                        else:
                            failed_periods.append(f"{periodo} ({folha_type}) - células já preenchidas")
                    else:
                        failed_periods.append(f"{periodo} ({folha_type}) - linha não encontrada")
            
            workbook.save(excel_path)
            
            # Resultado final
            success_periods = len(successful_periods)
            self._log(f"Processamento concluído: {success_periods}/{total_periods} períodos atualizados")
            
            if failed_periods:
                self._log(f"Falhas em {len(failed_periods)} períodos", "WARNING")
                for failed in failed_periods[:3]:
                    self._log(f"  {failed}", "WARNING")
                if len(failed_periods) > 3:
                    self._log(f"  ... e mais {len(failed_periods) - 3} períodos", "WARNING")
            
            return {
                'total_periods': total_periods,
                'success_periods': success_periods,
                'failed_periods': failed_periods,
                'updates_count': updates_count
            }
            
        except Exception as e:
            self._log(f"Erro ao atualizar Excel: {e}", "ERROR")
            raise

    def process_pdf(self, pdf_filename: str) -> Dict:
        """
        Processa PDF completo
        
        Returns:
            Dict com resultados do processamento
        """
        try:
            self._update_progress(0, "Iniciando processamento...")
            
            # Carrega configuração se não foi definida manualmente
            if not self.trabalho_dir:
                try:
                    self.load_env_config()
                except:
                    pass  # Será tratado na validação
            
            # Valida configuração
            valid, message = self.validate_trabalho_dir()
            if not valid:
                raise ValueError(message)
            
            # Encontra PDF
            pdf_path = self.find_pdf_file(pdf_filename)
            self._log(f"PDF encontrado: {Path(pdf_path).name}")
            
            # Extrai nome da pessoa
            self._update_progress(5, "Detectando nome da pessoa...")
            person_name = self.extract_person_name_from_pdf(pdf_path)
            
            if person_name:
                self._log(f"Nome detectado: {person_name}")
                excel_path = self.copy_modelo_to_dados(pdf_path, person_name)
                arquivo_final = f"DADOS/{self.normalize_filename(person_name)}.xlsm"
            else:
                self._log("Nome não detectado - usando nome do PDF")
                excel_path = self.copy_modelo_to_dados(pdf_path)
                arquivo_final = f"DADOS/{Path(pdf_path).stem}.xlsm"
            
            self._log(f"Arquivo criado: {arquivo_final}")
            
            # Extrai dados do PDF
            self._update_progress(10, "Extraindo texto do PDF...")
            pages_text = self.extract_text_from_pdf(pdf_path)
            
            # Categoriza páginas
            self._update_progress(30, "Categorizando páginas...")
            categorized_pages = self.filter_and_categorize_pages(pages_text)
            
            total_pages = len(pages_text)
            folha_normal_count = len(categorized_pages['FOLHA NORMAL'])
            salario_13_count = len(categorized_pages['13 SALARIO'])
            
            self._log(f"PDF processado: {total_pages} páginas totais")
            self._log(f"  - FOLHA NORMAL: {folha_normal_count} páginas")
            self._log(f"  - 13 SALARIO: {salario_13_count} páginas")
            
            # Extrai dados de cada tipo de página
            self._update_progress(40, "Extraindo dados das páginas...")
            extracted_data = {
                'FOLHA NORMAL': {},
                '13 SALARIO': {}
            }
            
            total_valid_pages = folha_normal_count + salario_13_count
            current_page = 0
            
            for folha_type, pages in categorized_pages.items():
                if not pages:
                    continue
                    
                for i, page_text in enumerate(pages):
                    current_page += 1
                    
                    # Progresso da extração de dados (40-70% do total)
                    progress = int(40 + (current_page / total_valid_pages) * 30)
                    self._update_progress(progress, f"Processando {folha_type} - página {i+1}")
                    
                    date_ref = self.extract_reference_date(page_text)
                    if not date_ref:
                        continue
                    
                    page_data = self.extract_data_from_page(page_text, folha_type)
                    
                    if page_data:
                        extracted_data[folha_type][date_ref] = page_data
            
            # Atualiza Excel
            total_extracted = sum(len(periods) for periods in extracted_data.values())
            if total_extracted > 0:
                self._update_progress(70, "Atualizando planilha Excel...")
                excel_results = self.update_excel_file(excel_path, extracted_data)
                
                self._update_progress(100, "Processamento concluído!")
                
                return {
                    'success': True,
                    'person_name': person_name,
                    'excel_path': excel_path,
                    'arquivo_final': arquivo_final,
                    'total_pages': total_pages,
                    'folha_normal_count': folha_normal_count,
                    'salario_13_count': salario_13_count,
                    'total_extracted': total_extracted,
                    'folha_normal_periods': len(extracted_data.get('FOLHA NORMAL', {})),
                    'salario_13_periods': len(extracted_data.get('13 SALARIO', {})),
                    **excel_results
                }
            else:
                raise ValueError("Nenhum dado foi extraído do PDF!")
                
        except Exception as e:
            self._log(f"Erro no processamento: {e}", "ERROR")
            self._update_progress(0, f"Erro: {e}")
            return {
                'success': False,
                'error': str(e)
            }