#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF para Excel Updater - Processamento com Diretório de Trabalho
================================================================

Versão 3.2 - Detecção de Nome + FOLHA NORMAL + 13 SALÁRIO

Funcionalidades:
- Diretório de trabalho obrigatório configurado via MODELO_DIR no .env
- Busca PDF e MODELO.xlsm no diretório de trabalho
- Cria pasta DADOS no diretório de trabalho
- Interface gráfica para seleção de PDF (opcional)
- Detecta nome da pessoa no PDF e usa para nomear arquivo final
- Processa FOLHA NORMAL (linhas 1-65) e 13 SALARIO (linhas 67+)
- Pode ser executado de qualquer local desde que .env esteja configurado

Estrutura esperada no diretório de trabalho:
├── MODELO.xlsm          # Planilha modelo (obrigatório)
├── arquivo.pdf          # PDF a processar
└── DADOS/               # Pasta criada automaticamente
    └── NOME DA PESSOA.xlsm     # Resultado com nome da pessoa

Configuração .env:
MODELO_DIR=C:\\caminho\\para\\diretorio\\de\\trabalho

Uso:
python pdf_to_excel_updater.py                    # Abre seletor de arquivo
python pdf_to_excel_updater.py arquivo.pdf        # Processa arquivo específico

Autor: Sistema de Extração Automatizada
Data: 2025
"""

import re
import pandas as pd
import pdfplumber
from datetime import datetime
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple
import argparse
from openpyxl import load_workbook
import os
import shutil
from dotenv import load_dotenv
import tkinter as tk
from tkinter import filedialog, messagebox

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PDFToExcelUpdater:
    """Classe para processar PDF usando diretório de trabalho configurado"""
    
    def __init__(self):
        """Inicializa o updater com as regras de mapeamento"""
        
        # Regras de mapeamento para colunas específicas do Excel
        self.mapping_rules = {
            # FOLHA NORMAL - Obter da coluna ÍNDICE (com fallback especial para PRODUÇÃO)
            '01003601': {'code': 'PREMIO PROD. MENSAL', 'excel_column': 'X', 'source': 'indice', 'fallback_to_valor': True, 'folha_type': 'FOLHA NORMAL'},
            '01007301': {'code': 'HORAS EXT.100%-180', 'excel_column': 'Y', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '01009001': {'code': 'ADIC.NOT.25%-180', 'excel_column': 'AC', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '01003501': {'code': 'HORAS EXT.75%-180', 'excel_column': 'AA', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            '02007501': {'code': 'DIFER.PROV. HORAS EXTRAS 75%', 'excel_column': 'AA', 'source': 'indice', 'folha_type': 'FOLHA NORMAL'},
            
            # FOLHA NORMAL - Obter da coluna VALOR
            '09090301_NORMAL': {'code': 'SALARIO CONTRIB INSS', 'excel_column': 'B', 'source': 'valor', 'folha_type': 'FOLHA NORMAL', 'original_code': '09090301'},
            
            # 13 SALARIO - Obter da coluna VALOR
            '09090301_13SAL': {'code': 'SALARIO CONTRIB INSS', 'excel_column': 'B', 'source': 'valor', 'folha_type': '13 SALARIO', 'original_code': '09090301'},
            '09090101_13SAL': {'code': 'REMUNERACAO BRUTA', 'excel_column': 'B', 'source': 'valor', 'folha_type': '13 SALARIO', 'original_code': '09090101', 'is_fallback': True}
        }
        
        # Planilha preferida (pode ser especificada externamente)
        self.preferred_sheet = None
        
        # Diretório de trabalho
        self.trabalho_dir = None
        
        # Meses em português para conversão
        self.meses_pt = {
            'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
            'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
            'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
        }
        
        # Abreviações de meses
        self.meses_abrev = {
            'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
        }

    def load_env_config(self):
        """Carrega configurações do arquivo .env"""
        # Carrega o arquivo .env se existir
        if Path('.env').exists():
            load_dotenv()
            logger.debug("Arquivo .env carregado")
        else:
            raise ValueError("Arquivo .env não encontrado. Configure MODELO_DIR no arquivo .env")
        
        # Obtém o diretório de trabalho
        self.trabalho_dir = os.getenv('MODELO_DIR')
        if not self.trabalho_dir:
            raise ValueError("MODELO_DIR não está definido no arquivo .env")
        
        # Valida se diretório existe
        trabalho_dir_path = Path(self.trabalho_dir)
        if not trabalho_dir_path.exists():
            raise ValueError(f"Diretório de trabalho não encontrado: {self.trabalho_dir}")
        
        logger.debug(f"Diretório de trabalho configurado: {self.trabalho_dir}")

    def select_pdf_file(self) -> Optional[str]:
        """Abre diálogo para seleção de arquivo PDF no diretório de trabalho"""
        try:
            # Carrega configuração primeiro
            self.load_env_config()
            
            # Cria janela invisível
            root = tk.Tk()
            root.withdraw()  # Esconde a janela principal
            root.attributes('-topmost', True)  # Mantém diálogo na frente
            
            # Configura diálogo para abrir no diretório de trabalho
            trabalho_dir_path = Path(self.trabalho_dir)
            
            # Procura arquivos PDF no diretório de trabalho
            pdf_files = list(trabalho_dir_path.glob("*.pdf"))
            
            if not pdf_files:
                messagebox.showwarning(
                    "Nenhum PDF encontrado", 
                    f"Nenhum arquivo PDF encontrado no diretório de trabalho:\n{self.trabalho_dir}"
                )
                root.destroy()
                return None
            
            # Abre diálogo de seleção
            selected_file = filedialog.askopenfilename(
                title="Selecione o arquivo PDF para processar",
                initialdir=str(trabalho_dir_path),
                filetypes=[
                    ("Arquivos PDF", "*.pdf"),
                    ("Todos os arquivos", "*.*")
                ]
            )
            
            root.destroy()
            
            if selected_file:
                # Retorna apenas o nome do arquivo (sem caminho)
                return Path(selected_file).name
            else:
                return None
                
        except Exception as e:
            if 'root' in locals():
                root.destroy()
            logger.error(f"Erro ao abrir diálogo de seleção: {e}")
            return None

    def extract_person_name_from_pdf(self, pdf_path: str) -> Optional[str]:
        """Extrai o nome da pessoa da primeira página do PDF"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                if not pdf.pages:
                    logger.debug("PDF não possui páginas")
                    return None
                
                # Extrai texto apenas da primeira página
                first_page = pdf.pages[0]
                text = first_page.extract_text()
                
                if not text:
                    logger.debug("Primeira página não possui texto extraível")
                    return None
                
                logger.debug("Procurando nome na primeira página do PDF...")
                
                # Procura por padrões de nome
                name_patterns = [
                    r'Nome\s*:\s*([A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ\s]+?)(?:\n|$|[A-Z]{2,}:)',
                    r'NOME\s*:\s*([A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ\s]+?)(?:\n|$|[A-Z]{2,}:)',
                    r'Nome\s*:\s*(.+?)(?:\n|Endereço|CPF|RG)',
                    r'NOME\s*:\s*(.+?)(?:\n|ENDEREÇO|CPF|RG)',
                    r'Nome\s*:\s*(.+?)$',
                    r'NOME\s*:\s*(.+?)$'
                ]
                
                lines = text.split('\n')
                
                # Debug: mostra as primeiras linhas para análise
                logger.debug("Primeiras 10 linhas da página:")
                for i, line in enumerate(lines[:10]):
                    line_clean = line.strip()
                    if line_clean:
                        logger.debug(f"  {i+1}: {line_clean}")
                
                # Procura linha por linha
                for line_num, line in enumerate(lines):
                    line_clean = line.strip()
                    
                    # Procura padrões na linha atual
                    for pattern_num, pattern in enumerate(name_patterns):
                        match = re.search(pattern, line_clean, re.IGNORECASE)
                        if match:
                            nome_bruto = match.group(1).strip()
                            
                            # Limpa e valida o nome
                            nome_limpo = self.clean_extracted_name(nome_bruto)
                            
                            if nome_limpo:
                                logger.debug(f"Nome encontrado (padrão {pattern_num+1}, linha {line_num+1}): '{nome_bruto}' -> '{nome_limpo}'")
                                return nome_limpo
                            else:
                                logger.debug(f"Nome descartado (padrão {pattern_num+1}, linha {line_num+1}): '{nome_bruto}' (inválido após limpeza)")
                
                logger.debug("Nenhum nome válido encontrado na primeira página")
                return None
                
        except Exception as e:
            logger.error(f"Erro ao extrair nome do PDF: {e}")
            return None

    def clean_extracted_name(self, nome_bruto: str) -> Optional[str]:
        """Limpa e valida o nome extraído"""
        if not nome_bruto:
            return None
        
        # Remove espaços extras e converte para maiúsculo
        nome = nome_bruto.strip().upper()
        
        # Remove caracteres não desejados
        nome = re.sub(r'[^\w\s]', ' ', nome)
        
        # Remove múltiplos espaços
        nome = re.sub(r'\s+', ' ', nome).strip()
        
        # Validações básicas
        if len(nome) < 3:  # Muito curto
            return None
        
        if len(nome) > 100:  # Muito longo
            return None
        
        # Verifica se não é só números
        if nome.replace(' ', '').isdigit():
            return None
        
        # Verifica se tem pelo menos algumas letras
        if not re.search(r'[A-ZÁÇÃÂÊÔÉÍÓÚÀÈÌÒÙ]', nome):
            return None
        
        # Remove palavras comuns que podem aparecer
        palavras_excluir = ['NOME', 'FUNCIONARIO', 'FUNCIONÁRIO', 'TRABALHADOR', 'COLABORADOR', 'EMPREGADO']
        palavras = nome.split()
        palavras_filtradas = [p for p in palavras if p not in palavras_excluir]
        
        if not palavras_filtradas:
            return None
        
        nome_final = ' '.join(palavras_filtradas)
        
        # Segunda validação
        if len(nome_final) < 3:
            return None
        
        return nome_final

    def normalize_filename(self, nome: str) -> str:
        """Converte nome da pessoa para formato de arquivo válido mantendo espaços"""
        # Mantém espaços como estão no PDF original
        filename = nome
        
        # Remove apenas caracteres que são realmente inválidos para nomes de arquivo
        # Windows: < > : " / \ | ? *
        filename = re.sub(r'[<>:"/\\|?*]', '', filename)
        
        # Remove apenas caracteres de controle perigosos, mas mantém letras, números e espaços
        filename = re.sub(r'[\x00-\x1f\x7f]', '', filename)
        
        # Remove espaços múltiplos consecutivos
        filename = re.sub(r'\s+', ' ', filename)
        
        # Remove espaços no início e fim
        filename = filename.strip()
        
        # Limita tamanho (nomes de arquivo muito longos podem dar problema)
        if len(filename) > 100:
            filename = filename[:100].rstrip()
        
        return filename

    def find_pdf_file(self, pdf_filename: str) -> str:
        """Procura arquivo PDF no diretório de trabalho"""
        trabalho_dir_path = Path(self.trabalho_dir)
        
        # Se pdf_filename já tem caminho, usa direto
        if Path(pdf_filename).is_absolute() or '/' in pdf_filename or '\\' in pdf_filename:
            if Path(pdf_filename).exists():
                return pdf_filename
            else:
                raise ValueError(f"Arquivo PDF não encontrado: {pdf_filename}")
        
        # Procura no diretório de trabalho
        pdf_path = trabalho_dir_path / pdf_filename
        if pdf_path.exists():
            logger.debug(f"PDF encontrado no diretório de trabalho: {pdf_path}")
            return str(pdf_path)
        
        # Tenta com extensão .pdf se não foi fornecida
        if not pdf_filename.lower().endswith('.pdf'):
            pdf_path_with_ext = trabalho_dir_path / f"{pdf_filename}.pdf"
            if pdf_path_with_ext.exists():
                logger.debug(f"PDF encontrado (com extensão): {pdf_path_with_ext}")
                return str(pdf_path_with_ext)
        
        raise ValueError(f"Arquivo PDF '{pdf_filename}' não encontrado no diretório de trabalho: {self.trabalho_dir}")

    def copy_modelo_to_dados(self, pdf_path: str, custom_name: Optional[str] = None) -> str:
        """Copia o arquivo modelo para a pasta DADOS com nome personalizado ou baseado no PDF"""
        
        trabalho_dir_path = Path(self.trabalho_dir)
        
        # Procura pelo arquivo MODELO.xlsm no diretório de trabalho
        modelo_file = trabalho_dir_path / "MODELO.xlsm"
        if not modelo_file.exists():
            raise ValueError(f"Arquivo MODELO.xlsm não encontrado no diretório de trabalho: {self.trabalho_dir}")
        
        logger.debug(f"Arquivo modelo encontrado: {modelo_file}")
        
        # Cria pasta DADOS no diretório de trabalho
        dados_dir = trabalho_dir_path / "DADOS"
        dados_dir.mkdir(exist_ok=True)
        logger.debug(f"Pasta DADOS: {dados_dir}")
        
        # Define nome do arquivo de destino
        if custom_name:
            # Usa nome da pessoa detectado
            base_name = self.normalize_filename(custom_name)
            logger.debug(f"Usando nome personalizado: '{custom_name}' -> '{base_name}'")
        else:
            # Usa nome baseado no PDF como fallback
            base_name = Path(pdf_path).stem
            logger.debug(f"Usando nome baseado no PDF: '{base_name}'")
        
        destino_file = dados_dir / f"{base_name}.xlsm"
        
        # Copia o modelo para o destino
        try:
            shutil.copy2(modelo_file, destino_file)
            logger.debug(f"Modelo copiado: {modelo_file} -> {destino_file}")
            return str(destino_file)
        except Exception as e:
            raise ValueError(f"Erro ao copiar modelo: {e}")

    def extract_text_from_pdf(self, pdf_path: str) -> List[str]:
        """Extrai texto de todas as páginas do PDF"""
        pages_text = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                logger.debug(f"Processando PDF: {pdf_path} ({len(pdf.pages)} páginas)")
                
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
                    
        except Exception as e:
            logger.error(f"Erro ao processar PDF: {e}")
            raise
            
        return pages_text

    def extract_reference_date(self, text: str) -> Optional[Tuple[int, int]]:
        """Extrai a data de referência da página (mês/ano)"""
        patterns = [
            r'Referência:\s*(\w+)/(\d{4})',
            r'Referencia:\s*(\w+)/(\d{4})',
            r'Data\s*do\s*c[aá]lculo:\s*\d{2}/(\d{2})/(\d{4})',
            r'Per[ií]odo:\s*(\w+)/(\d{4})',
            r'Compet[êe]ncia:\s*(\w+)/(\d{4})',
            r'(\w+)\s*/\s*(\d{4})',
        ]
        
        found_matches = []
        
        for i, pattern in enumerate(patterns):
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                found_matches.append((i, match))
                try:
                    if len(match) == 2:
                        mes_str = match[0].lower()
                        ano = int(match[1])
                        
                        # Tenta converter mês por nome
                        mes = self.meses_pt.get(mes_str) or self.meses_abrev.get(mes_str)
                        
                        # Se não conseguir por nome, tenta por número
                        if not mes:
                            try:
                                mes = int(mes_str)
                                if 1 <= mes <= 12:
                                    logger.debug(f"Data extraída (padrão {i+1}): {mes_str}/{ano} -> {mes}/{ano}")
                                    return (mes, ano)
                            except ValueError:
                                continue
                        else:
                            logger.debug(f"Data extraída (padrão {i+1}): {mes_str}/{ano} -> {mes}/{ano}")
                            return (mes, ano)
                except ValueError:
                    continue
        
        # Log quando não conseguir extrair data
        if found_matches:
            logger.debug(f"Padrões encontrados mas inválidos: {found_matches}")
        else:
            logger.debug("Nenhum padrão de data encontrado na página")
            # Mostra as primeiras linhas para debug
            lines = text.split('\n')[:5]
            logger.debug(f"Primeiras linhas da página: {[line.strip() for line in lines if line.strip()]}")
        
        return None

    def extract_last_two_numbers(self, line: str):
        """Extrai os dois últimos números de uma linha, ignorando campos fantasma"""
        
        def convert_to_float_robust(value_str):
            if not value_str or not value_str.strip():
                return None
                
            cleaned = value_str.strip()
            
            # Detecta formato de horas (06:34) e converte para (06,34)
            if ':' in cleaned:
                hour_pattern = r'^\d{1,2}:\d{2}$'
                if re.match(hour_pattern, cleaned):
                    logger.debug(f"Formato de horas detectado: {cleaned}")
                    return cleaned.replace(':', ',')
            
            # Remove caracteres não numéricos exceto ponto e vírgula
            cleaned = re.sub(r'[^\d.,]', '', cleaned)
            
            if not cleaned:
                return None
            
            # Estratégias de conversão para números
            try:
                # Formato brasileiro (1.234,56)
                if ',' in cleaned and cleaned.count(',') == 1:
                    return float(cleaned.replace('.', '').replace(',', '.'))
                # Formato americano (1,234.56)
                elif '.' in cleaned and cleaned.count('.') == 1 and ',' in cleaned:
                    return float(cleaned.replace(',', ''))
                # Apenas vírgula decimal (1234,56)
                elif ',' in cleaned and '.' not in cleaned:
                    return float(cleaned.replace(',', '.'))
                # Apenas ponto decimal (1234.56)
                elif '.' in cleaned and ',' not in cleaned:
                    return float(cleaned)
                else:
                    return float(cleaned)
            except ValueError:
                return None
        
        # Encontra todos os números na linha (incluindo formato de horas)
        number_pattern = r'[\d]+(?:[.,:]\d+)*'
        matches = re.findall(number_pattern, line)
        
        if len(matches) >= 2:
            # Pega os dois últimos números
            penultimo = convert_to_float_robust(matches[-2])
            ultimo = convert_to_float_robust(matches[-1])
            return penultimo, ultimo
        elif len(matches) == 1:
            # Só um número - considera como valor
            ultimo = convert_to_float_robust(matches[-1])
            return None, ultimo
        else:
            return None, None

    def extract_data_from_page(self, text: str, folha_type: str) -> Dict[str, any]:
        """Extrai dados específicos de uma página usando as regras de mapeamento"""
        data = {}
        codes_found = []
        
        # Filtra regras pelo tipo de folha
        relevant_rules = {k: v for k, v in self.mapping_rules.items() if v.get('folha_type') == folha_type}
        
        logger.debug(f"Processando página como: {folha_type}")
        logger.debug(f"Regras aplicáveis: {list(relevant_rules.keys())}")
        
        # Processa linha por linha
        lines = text.split('\n')
        
        # Para 13 SALARIO, precisamos tratar fallback especial entre 09090301 e 09090101
        found_09090301 = None
        found_09090101 = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Verifica códigos específicos do tipo de folha
            for rule_key, rule in relevant_rules.items():
                original_code = rule.get('original_code', rule_key)
                
                if original_code in line:
                    codes_found.append(original_code)
                    
                    # Extrai os dois últimos números da linha
                    indice, valor = self.extract_last_two_numbers(line)
                    
                    logger.debug(f"Processando {original_code} ({folha_type}): índice={indice}, valor={valor}")
                    
                    # Para 13 SALARIO, armazena os valores encontrados para fallback
                    if folha_type == '13 SALARIO':
                        if original_code == '09090301':
                            found_09090301 = valor
                        elif original_code == '09090101':
                            found_09090101 = valor
                    
                    # Lógica de mapeamento normal para outros códigos
                    if folha_type == 'FOLHA NORMAL' or (folha_type == '13 SALARIO' and original_code not in ['09090301', '09090101']):
                        value_to_use = None
                        source_used = None
                        
                        if rule['source'] == 'indice':
                            if indice is not None and indice != 0:
                                value_to_use = indice
                                source_used = 'índice'
                            elif rule.get('fallback_to_valor', False) and valor is not None:
                                # Fallback para PRODUÇÃO (01003601)
                                value_to_use = valor
                                source_used = 'valor (fallback)'
                                logger.debug(f"[FALLBACK] {original_code}: Usando fallback valor pois indice vazio")
                        elif rule['source'] == 'valor' and valor is not None:
                            value_to_use = valor
                            source_used = 'valor'
                        
                        if value_to_use is not None:
                            data[rule['excel_column']] = value_to_use
                            logger.debug(f"[OK] {original_code}: Coluna {rule['excel_column']} = {value_to_use} ({source_used})")
                        else:
                            logger.debug(f"[AVISO] {original_code}: Valor não pôde ser extraído")
        
        # Para 13 SALARIO, aplica lógica de fallback entre 09090301 e 09090101
        if folha_type == '13 SALARIO':
            value_to_use = None
            source_used = None
            
            if found_09090301 is not None and found_09090301 != 0:
                value_to_use = found_09090301
                source_used = '09090301 (prioridade)'
            elif found_09090101 is not None and found_09090101 != 0:
                value_to_use = found_09090101
                source_used = '09090101 (fallback)'
                logger.debug(f"[FALLBACK 13SAL] Usando 09090101 pois 09090301 não disponível")
            
            if value_to_use is not None:
                data['B'] = value_to_use
                logger.debug(f"[OK] 13 SALARIO: Coluna B = {value_to_use} ({source_used})")
            else:
                if found_09090301 is not None or found_09090101 is not None:
                    logger.debug(f"[AVISO] 13 SALARIO: Códigos encontrados mas valores zerados (09090301={found_09090301}, 09090101={found_09090101})")
        
        if codes_found:
            logger.debug(f"Códigos encontrados na página: {codes_found}")
        else:
            logger.debug("Nenhum código conhecido encontrado na página")
            # Mostra algumas linhas para debug se não encontrar nada
            sample_lines = [line for line in lines if line.strip() and not line.startswith(' ')][:3]
            if sample_lines:
                logger.debug(f"Amostra de linhas: {sample_lines}")
        
        return data

    def filter_and_categorize_pages(self, pages_text: List[str]) -> Dict[str, List[str]]:
        """Filtra e categoriza páginas por tipo (FOLHA NORMAL e 13 SALARIO)"""
        categorized_pages = {
            'FOLHA NORMAL': [],
            '13 SALARIO': []
        }
        
        for i, text in enumerate(pages_text):
            page_type = None
            page_type_found = False
            
            # Primeiro, procura especificamente pela linha "Tipo da folha:"
            lines = text.split('\n')
            for line in lines:
                line_clean = line.strip()
                
                # Verifica se é uma linha de tipo de folha
                if re.search(r'Tipo\s+da\s+folha\s*:', line_clean, re.IGNORECASE):
                    page_type_found = True
                    logger.debug(f"Página {i+1}: Linha tipo encontrada: '{line_clean}'")
                    
                    # Verifica se é folha normal
                    if re.search(r'FOLHA\s+NORMAL', line_clean, re.IGNORECASE):
                        page_type = 'FOLHA NORMAL'
                        logger.debug(f"Página {i+1}: FOLHA NORMAL identificada")
                        break
                    # Verifica se é 13º salário (sem º)
                    elif re.search(r'13\s*SAL[AÁ]RIO', line_clean, re.IGNORECASE):
                        page_type = '13 SALARIO'
                        logger.debug(f"Página {i+1}: 13 SALARIO identificada")
                        break
                    # Verifica se é outros tipos especiais (férias, rescisão, etc.) - IGNORAR
                    elif re.search(r'F[ÉE]RIAS|ADIANTAMENTO|RESCIS[ÃA]O', line_clean, re.IGNORECASE):
                        page_type = 'IGNORAR'
                        logger.debug(f"Página {i+1}: Tipo especial ignorado: '{line_clean}'")
                        break
            
            # Se não encontrou linha "Tipo da folha:", usa filtro de fallback mais restritivo
            if not page_type_found:
                logger.debug(f"Página {i+1}: Linha 'Tipo da folha' não encontrada, aplicando filtro de fallback")
                
                # Procura por indicadores de folhas especiais no cabeçalho/início da página
                header_text = '\n'.join(lines[:10])  # Apenas primeiras 10 linhas
                
                # Verifica 13º salário no cabeçalho
                if re.search(r'13\s*SAL[AÁ]RIO', header_text, re.IGNORECASE):
                    page_type = '13 SALARIO'
                    logger.debug(f"Página {i+1}: 13 SALARIO identificado no cabeçalho (fallback)")
                # Verifica outros tipos para ignorar
                elif re.search(r'F[ÉE]RIAS|ADIANTAMENTO\s*SALARIAL|RESCIS[ÃA]O', header_text, re.IGNORECASE):
                    page_type = 'IGNORAR'
                    logger.debug(f"Página {i+1}: Tipo especial identificado no cabeçalho (fallback)")
                else:
                    # Se não detectou nada específico, assume folha normal
                    page_type = 'FOLHA NORMAL'
                    logger.debug(f"Página {i+1}: Assumindo FOLHA NORMAL (fallback)")
            
            # Adiciona à categoria apropriada
            if page_type and page_type != 'IGNORAR':
                categorized_pages[page_type].append(text)
                logger.debug(f"Página {i+1}: Aceita como {page_type}")
            else:
                logger.debug(f"Página {i+1}: Rejeitada (ignorada)")
                
        logger.debug(f"Páginas categorizadas: FOLHA NORMAL={len(categorized_pages['FOLHA NORMAL'])}, 13 SALARIO={len(categorized_pages['13 SALARIO'])}")
        return categorized_pages

    def find_row_for_period(self, worksheet, month: int, year: int, folha_type: str) -> Optional[int]:
        """Encontra a linha correspondente ao período (mês/ano) na planilha"""
        meses_nomes = ['', 'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                      'jul', 'ago', 'set', 'out', 'nov', 'dez']
        periodo_procurado = f"{meses_nomes[month]}/{str(year)[2:]}"
        
        # Define faixa de linhas baseada no tipo de folha
        if folha_type == 'FOLHA NORMAL':
            start_row = 1
            end_row = 65  # Até linha 65 para folha normal
        elif folha_type == '13 SALARIO':
            start_row = 67  # A partir da linha 67 (66 é cabeçalho)
            end_row = worksheet.max_row
        else:
            logger.error(f"Tipo de folha não reconhecido: {folha_type}")
            return None
        
        logger.debug(f"Procurando período: {periodo_procurado} (mês {month}, ano {year}) - {folha_type} (linhas {start_row}-{end_row})")
        
        found_values = []  # Para debug
        
        # Procura na coluna A (PERÍODO) na faixa específica
        for row_num in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
            cell_value = worksheet[f'A{row_num}'].value
            
            if cell_value is None:
                continue
            
            # Armazena valores encontrados para debug (apenas primeiros 10)
            if len(found_values) < 10:
                found_values.append(f"A{row_num}: {repr(cell_value)}")
                
            # Se for string, tenta formato texto (nov/12, dez/12, etc.)
            if isinstance(cell_value, str):
                cell_str = cell_value.strip()
                
                if cell_str == periodo_procurado:
                    logger.debug(f"[OK] Período {periodo_procurado} encontrado na linha {row_num} (formato texto) - {folha_type}")
                    return row_num
            
            # Se for data/datetime, compara mês e ano
            elif isinstance(cell_value, datetime):
                if cell_value.month == month and cell_value.year == year:
                    logger.debug(f"[OK] Período {periodo_procurado} encontrado na linha {row_num} (formato data: {cell_value}) - {folha_type}")
                    return row_num
            
            # Se for number (serial date do Excel), tenta converter
            elif isinstance(cell_value, (int, float)):
                try:
                    from datetime import timedelta
                    
                    # Excel epoch é 1900-01-01, mas com bug: conta 1900 como bissexto
                    if cell_value > 59:  # Após 28/02/1900
                        excel_date = datetime(1899, 12, 30) + timedelta(days=cell_value)
                    else:
                        excel_date = datetime(1899, 12, 31) + timedelta(days=cell_value)
                    
                    if excel_date.month == month and excel_date.year == year:
                        logger.debug(f"[OK] Período {periodo_procurado} encontrado na linha {row_num} (serial: {cell_value} -> {excel_date}) - {folha_type}")
                        return row_num
                        
                except Exception as e:
                    logger.debug(f"Erro ao converter serial date {cell_value}: {e}")
                    continue
        
        # Log de debug quando não encontrar
        logger.warning(f"[ERRO] Período {periodo_procurado} não encontrado na planilha - {folha_type} (linhas {start_row}-{end_row})")
        if found_values:
            logger.debug(f"Valores encontrados na coluna A: {', '.join(found_values[:5])}")
            if len(found_values) > 5:
                logger.debug(f"... e mais {len(found_values) - 5} valores")
        else:
            logger.debug(f"Coluna A vazia ou sem dados válidos na faixa {start_row}-{end_row}")
        
        return None

    def update_excel_file(self, excel_path: str, extracted_data: Dict):
        """Atualiza o arquivo Excel existente com os dados extraídos"""
        try:
            # Verifica se é arquivo com macros (.xlsm)
            is_macro_enabled = excel_path.lower().endswith('.xlsm')
            
            # Carrega a planilha existente preservando macros se necessário
            if is_macro_enabled:
                workbook = load_workbook(excel_path, keep_vba=True)
            else:
                workbook = load_workbook(excel_path)
            
            worksheet = None
            
            # Se foi especificada uma planilha, tenta usar ela primeiro
            if self.preferred_sheet:
                if self.preferred_sheet in workbook.sheetnames:
                    worksheet = workbook[self.preferred_sheet]
                    logger.debug(f"Usando planilha especificada: {worksheet.title}")
                else:
                    raise ValueError(f"Planilha especificada '{self.preferred_sheet}' não encontrada")
            else:
                # Regra: "LEVANTAMENTO DADOS" é obrigatória se não especificada
                if 'LEVANTAMENTO DADOS' in workbook.sheetnames:
                    worksheet = workbook['LEVANTAMENTO DADOS']
                    logger.debug(f"Usando planilha padrão: {worksheet.title}")
                else:
                    raise ValueError("Planilha 'LEVANTAMENTO DADOS' não encontrada. Use -s para especificar outra planilha.")
            
            updates_count = 0
            failed_periods = []
            successful_periods = []
            
            # Separar dados por tipo de folha
            for folha_type in ['FOLHA NORMAL', '13 SALARIO']:
                if folha_type not in extracted_data:
                    continue
                    
                logger.debug(f"\n--- Processando {folha_type} ---")
                
                # Para cada período com dados extraídos deste tipo
                for (month, year), data in extracted_data[folha_type].items():
                    meses = ['', 'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                            'jul', 'ago', 'set', 'out', 'nov', 'dez']
                    periodo = f"{meses[month]}/{str(year)[2:]}"
                    
                    # Encontra a linha correspondente ao período na faixa correta
                    row_num = self.find_row_for_period(worksheet, month, year, folha_type)
                    
                    if row_num:
                        # Atualiza cada coluna com dados
                        period_updates = 0
                        cells_already_filled = []
                        cells_updated = []
                        
                        for column, value in data.items():
                            cell_address = f"{column}{row_num}"
                            old_value = worksheet[cell_address].value
                            
                            # Só atualiza se a célula estiver vazia ou diferente
                            if old_value is None or old_value == '' or old_value == 0:
                                worksheet[cell_address] = value
                                cells_updated.append(f"{cell_address}:{value}")
                                logger.debug(f"Atualizado {cell_address}: {old_value} -> {value} ({folha_type})")
                                updates_count += 1
                                period_updates += 1
                            else:
                                cells_already_filled.append(f"{cell_address}:{old_value}")
                                logger.debug(f"Celula {cell_address} ja preenchida: {old_value} ({folha_type})")
                        
                        if period_updates > 0:
                            successful_periods.append(f"{periodo} ({folha_type})")
                            logger.debug(f"[OK] {periodo} ({folha_type}): {period_updates} células atualizadas")
                        else:
                            # Período encontrado mas nenhuma célula foi atualizada
                            failed_periods.append(f"{periodo} ({folha_type}) - células já preenchidas")
                            logger.debug(f"[AVISO] {periodo} ({folha_type}): todas as células já estavam preenchidas")
                    else:
                        # Período não encontrado na planilha
                        failed_periods.append(f"{periodo} ({folha_type}) - linha não encontrada")
                        logger.debug(f"[ERRO] {periodo} ({folha_type}): linha não encontrada na planilha")
            
            # Salva as alterações preservando formato original
            workbook.save(excel_path)
            
            # Log resumido
            total_periods = sum(len(periods) for periods in extracted_data.values())
            success_periods = len(successful_periods)
            
            if success_periods == total_periods:
                logger.info(f"[OK] Processamento concluído: {success_periods} períodos atualizados")
            else:
                logger.info(f"[AVISO] Processamento concluído: {success_periods}/{total_periods} períodos atualizados")
            
            if failed_periods:
                logger.warning(f"[ERRO] Falhas em {len(failed_periods)} períodos:")
                for failed in failed_periods[:3]:  # Mostra apenas primeiros 3
                    logger.warning(f"   {failed}")
                if len(failed_periods) > 3:
                    logger.warning(f"   ... e mais {len(failed_periods) - 3} períodos")
            
            logger.debug(f"Total de células atualizadas: {updates_count}")
            
        except Exception as e:
            logger.error(f"Erro ao atualizar Excel: {e}")
            raise

    def process_pdf(self, pdf_filename: str):
        """Processa PDF usando diretório de trabalho configurado"""
        logger.debug("Iniciando processamento PDF...")
        
        # Carrega configuração obrigatória
        self.load_env_config()
        
        # Encontra PDF no diretório de trabalho
        pdf_path = self.find_pdf_file(pdf_filename)
        logger.info(f"PDF encontrado: {Path(pdf_path).name}")
        
        # NOVO: Tenta extrair nome da pessoa do PDF
        person_name = self.extract_person_name_from_pdf(pdf_path)
        
        if person_name:
            logger.info(f"Nome detectado: {person_name}")
            # Copia modelo e cria arquivo com nome da pessoa
            excel_path = self.copy_modelo_to_dados(pdf_path, person_name)
            arquivo_final = f"DADOS/{self.normalize_filename(person_name)}.xlsm"
        else:
            logger.info("Nome não detectado - usando nome do PDF")
            # Copia modelo com nome baseado no PDF (comportamento anterior)
            excel_path = self.copy_modelo_to_dados(pdf_path)
            arquivo_final = f"DADOS/{Path(pdf_path).stem}.xlsm"
        
        logger.info(f"Arquivo criado: {arquivo_final}")
        
        # Extrai dados do PDF
        pages_text = self.extract_text_from_pdf(pdf_path)
        categorized_pages = self.filter_and_categorize_pages(pages_text)
        
        total_pages = len(pages_text)
        folha_normal_count = len(categorized_pages['FOLHA NORMAL'])
        salario_13_count = len(categorized_pages['13 SALARIO'])
        
        logger.debug(f"PDF processado: {total_pages} páginas totais")
        logger.debug(f"  - FOLHA NORMAL: {folha_normal_count} páginas")
        logger.debug(f"  - 13 SALARIO: {salario_13_count} páginas")
        
        # Extrai dados de cada tipo de página
        extracted_data = {
            'FOLHA NORMAL': {},
            '13 SALARIO': {}
        }
        
        for folha_type, pages in categorized_pages.items():
            if not pages:
                continue
                
            logger.debug(f"\n--- Processando páginas de {folha_type} ---")
            
            pages_sem_data = []
            pages_sem_periodo = []
            
            for i, page_text in enumerate(pages):
                logger.debug(f"\n--- Processando página {i+1} de {folha_type} ---")
                
                # Extrai data de referência
                date_ref = self.extract_reference_date(page_text)
                if not date_ref:
                    pages_sem_periodo.append(i+1)
                    logger.debug(f"[AVISO] Página {i+1} ({folha_type}): Data de referência não encontrada")
                    continue
                
                mes, ano = date_ref
                meses_nomes = ['', 'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                              'jul', 'ago', 'set', 'out', 'nov', 'dez']
                periodo_str = f"{meses_nomes[mes]}/{str(ano)[2:]}"
                logger.debug(f"Data: Página {i+1} ({folha_type}): Período identificado como {periodo_str}")
                
                # Extrai dados da página
                page_data = self.extract_data_from_page(page_text, folha_type)
                
                if page_data:
                    extracted_data[folha_type][date_ref] = page_data
                    logger.debug(f"[OK] Página {i+1} ({folha_type}): {len(page_data)} campos extraídos {list(page_data.keys())}")
                else:
                    pages_sem_data.append((i+1, periodo_str))
                    logger.debug(f"[AVISO] Página {i+1} ({folha_type}): Período {periodo_str} identificado, mas nenhum código mapeado encontrado")
            
            # Resumo do processamento das páginas por tipo
            if pages_sem_periodo:
                logger.info(f"[AVISO] {folha_type}: {len(pages_sem_periodo)} páginas sem período identificado: {pages_sem_periodo}")
            
            if pages_sem_data:
                periodos_sem_data = [periodo for _, periodo in pages_sem_data]
                logger.info(f"[AVISO] {folha_type}: {len(pages_sem_data)} páginas com período mas sem dados: {periodos_sem_data}")
        
        # Atualiza Excel
        total_extracted = sum(len(periods) for periods in extracted_data.values())
        if total_extracted > 0:
            logger.debug(f"\n--- Atualizando Excel ---")
            logger.debug(f"Períodos a processar: FOLHA NORMAL={len(extracted_data['FOLHA NORMAL'])}, 13 SALARIO={len(extracted_data['13 SALARIO'])}")
            self.update_excel_file(excel_path, extracted_data)
        else:
            logger.warning("ERRO: Nenhum dado foi extraído do PDF!")
            logger.warning("Possíveis causas: formato de data não reconhecido ou códigos não encontrados")
        
        return extracted_data, excel_path, person_name

def main():
    """Função principal da aplicação"""
    parser = argparse.ArgumentParser(description='PDF para Excel Updater v3.2 - Detecção de Nome + FOLHA NORMAL + 13 SALARIO')
    parser.add_argument('pdf_filename', nargs='?', help='Nome do arquivo PDF (opcional - abrirá diálogo se não fornecido)')
    parser.add_argument('-s', '--sheet', help='Nome da planilha específica (padrão: "LEVANTAMENTO DADOS")')
    parser.add_argument('-v', '--verbose', action='store_true', help='Modo verboso')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    else:
        # Modo silencioso - apenas INFO e WARNING/ERROR
        logging.getLogger().setLevel(logging.INFO)
    
    try:
        # Cria updater
        updater = PDFToExcelUpdater()
        
        # Se especificou planilha, usa diretamente
        if args.sheet:
            updater.preferred_sheet = args.sheet
        
        # Determina qual PDF processar
        pdf_filename = args.pdf_filename
        
        if not pdf_filename:
            # Se não foi fornecido PDF, abre diálogo de seleção
            print("Abrindo seletor de arquivo...")
            pdf_filename = updater.select_pdf_file()
            
            if not pdf_filename:
                print("CANCELADO: Nenhum arquivo selecionado")
                return 0
        
        print(f"Processando: {pdf_filename}")
        
        extracted_data, excel_path, person_name = updater.process_pdf(pdf_filename)
        
        # Resumo final
        total_extracted = sum(len(periods) for periods in extracted_data.values())
        if total_extracted > 0:
            folha_normal_periods = len(extracted_data.get('FOLHA NORMAL', {}))
            salario_13_periods = len(extracted_data.get('13 SALARIO', {}))
            
            print(f"OK: Concluído: {total_extracted} períodos processados")
            if folha_normal_periods > 0:
                print(f"  - FOLHA NORMAL: {folha_normal_periods} períodos")
            if salario_13_periods > 0:
                print(f"  - 13 SALARIO: {salario_13_periods} períodos")
            
            if person_name:
                excel_name = f"{self.normalize_filename(person_name)}.xlsm"
                print(f"Nome detectado: {person_name}")
            else:
                excel_name = Path(excel_path).name
                print("Nome não detectado - usando nome do PDF")
                
            print(f"Arquivo criado: DADOS/{excel_name}")
        else:
            print("ERRO: Nenhum dado foi extraído do PDF!")
            print("Dica: Use -v para diagnóstico detalhado")
            return 1
        
        return 0
        
    except ValueError as e:
        print(f"ERRO: {e}")
        return 1
    except Exception as e:
        print(f"ERRO: Erro inesperado: {e}")
        if args.verbose:
            raise
        return 1

if __name__ == "__main__":
    import sys
    sys.exit(main())